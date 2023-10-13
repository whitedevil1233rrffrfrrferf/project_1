from flask import  Flask,render_template,request,redirect
from flask_sqlalchemy import SQLAlchemy
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
app=Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI']="sqlite:///user.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"]=False
app.config['SECRET_KEY'] = 'login_system'
db=SQLAlchemy(app)

class Employer(db.Model):
    sno = db.Column(db.Integer, primary_key=True, autoincrement=True)   
    name=db.Column(db.String(200),nullable=False)
    age=db.Column(db.Integer,nullable=False)
    role=db.Column(db.String(200),nullable=False)

def extract_data_from_excel():
    wb=load_workbook("roles.xlsx")
    ws=wb.active
    for row in ws.iter_rows(min_row=5,values_only=True):
        if None not in row:
            sno,name,age,role=row
            existing_record=Employer.query.filter_by(name=name,role=role).first()
            if not existing_record:
                employer=Employer(name=name,age=age,role=role)
                db.session.add(employer)
    db.session.commit()        
@app.route("/",methods=["GET","POST"])
def home():
    if request.method=="POST":
        name=request.form["name"]
        age=request.form["age"]
        role=request.form["role"]
        existing_data=Employer.query.filter_by(name=name,age=age,role=role).first()
        if not existing_data:
            employee=Employer(name=name,age=age,role=role)
            db.session.add(employee)
            db.session.commit()
    data=Employer.query.all()        
    return render_template("index.html",data=data)

@app.route("/update/<int:sno>",methods=["GET","POST"])
def update(sno):
    if request.method=="POST":
        name=request.form["name"]
        age=request.form["age"]
        role=request.form["role"]
        user=Employer.query.filter_by(sno=sno).first()
        user.name=name
        user.role=role
        user.age=age
        db.session.add(user)
        db.session.commit()
        return redirect("/")
    user=Employer.query.filter_by(sno=sno).first()
    return render_template("update.html",user=user)
@app.route("/delete/<int:sno>")
def delete(sno):
    user=Employer.query.filter_by(sno=sno).first()
    db.session.delete(user)
    db.session.commit()
    return redirect("/")
with app.app_context():
        db.create_all()
        data=extract_data_from_excel()
         
if "__name__" == "__main__":
    
    app.run(debug=True)   
